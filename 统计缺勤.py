#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""统计超星学习通签到表中的缺勤次数。"""

from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_DIR = Path("算法设计与分析")
OUTPUT_FILE = Path("缺勤统计结果.xlsx")
ABSENT_STATUS = "未参与"


def find_header_row(ws) -> int | None:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and "姓名" in row and "签到状态" in row:
            return row_idx
    return None


def date_from_sheet(ws, file_path: Path) -> str:
    first_cell = ws["A1"].value
    if first_cell:
        text = str(first_cell).strip()
        if re.fullmatch(r"\d{8}", text):
            return text

    match = re.search(r"(20\d{6})", file_path.stem)
    return match.group(1) if match else file_path.stem


def read_attendance_file(file_path: Path) -> tuple[str, list[dict], Counter]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws)
    if header_row is None:
        return date_from_sheet(ws, file_path), [], Counter()

    headers = [cell.value for cell in ws[header_row]]
    col = {name: idx for idx, name in enumerate(headers) if name}
    required = ["姓名", "学号/工号", "行政班级", "签到状态"]
    if any(name not in col for name in required):
        return date_from_sheet(ws, file_path), [], Counter()

    date_text = date_from_sheet(ws, file_path)
    absent_rows: list[dict] = []
    status_counter: Counter = Counter()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue

        status = row[col["签到状态"]]
        status_counter[str(status or "").strip()] += 1

        if status == ABSENT_STATUS:
            absent_rows.append(
                {
                    "日期": date_text,
                    "姓名": row[col["姓名"]],
                    "学号": row[col["学号/工号"]],
                    "行政班级": row[col["行政班级"]],
                    "签到状态": status,
                    "来源文件": file_path.name,
                }
            )

    return date_text, absent_rows, status_counter


def autosize(ws, max_width: int = 45) -> None:
    for column_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), max_width)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def write_results(
    attendance_count: int,
    summary_rows: list[dict],
    detail_rows: list[dict],
    overview_rows: list[dict],
) -> None:
    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "缺勤次数统计"
    ws_summary.append(["排名", "姓名", "学号", "行政班级", "缺勤次数", "缺勤日期"])
    for rank, row in enumerate(summary_rows, start=1):
        ws_summary.append(
            [
                rank,
                row["姓名"],
                row["学号"],
                row["行政班级"],
                row["缺勤次数"],
                row["缺勤日期"],
            ]
        )
    style_sheet(ws_summary)

    ws_detail = wb.create_sheet("各次缺勤名单")
    ws_detail.append(["日期", "未参与人数", "姓名", "学号", "行政班级", "签到状态", "来源文件"])
    absent_by_date = Counter(row["日期"] for row in detail_rows)
    for row in detail_rows:
        ws_detail.append(
            [
                row["日期"],
                absent_by_date[row["日期"]],
                row["姓名"],
                row["学号"],
                row["行政班级"],
                row["签到状态"],
                row["来源文件"],
            ]
        )
    style_sheet(ws_detail)

    ws_overview = wb.create_sheet("签到概览")
    ws_overview.append(["总签到次数", attendance_count])
    ws_overview.append(["有缺勤记录学生数", len(summary_rows)])
    ws_overview.append([])
    ws_overview.append(["日期", "总人数", "已签", "教师代签", "病假", "未参与"])
    for row in overview_rows:
        ws_overview.append(
            [
                row["日期"],
                row["总人数"],
                row.get("已签", 0),
                row.get("教师代签", 0),
                row.get("病假", 0),
                row.get("未参与", 0),
            ]
        )
    for cell in ws_overview[1] + ws_overview[2] + ws_overview[4]:
        cell.font = Font(bold=True)
        if cell.row == 4:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")
    ws_overview.freeze_panes = "A5"
    ws_overview.auto_filter.ref = f"A4:F{ws_overview.max_row}"
    autosize(ws_overview)

    wb.save(OUTPUT_FILE)


def main() -> None:
    excel_files = sorted(SOURCE_DIR.glob("*.xlsx"))
    if not excel_files:
        print(f"未找到 Excel 文件：{SOURCE_DIR}")
        return

    student_info: dict[tuple[str, str], dict] = {}
    absent_dates: dict[tuple[str, str], list[str]] = defaultdict(list)
    detail_rows: list[dict] = []
    overview_rows: list[dict] = []

    for file_path in excel_files:
        date_text, absent_rows, status_counter = read_attendance_file(file_path)
        detail_rows.extend(absent_rows)

        for row in absent_rows:
            key = (str(row["姓名"]), str(row["学号"]))
            student_info[key] = {
                "姓名": row["姓名"],
                "学号": row["学号"],
                "行政班级": row["行政班级"],
            }
            absent_dates[key].append(row["日期"])

        overview_rows.append(
            {
                "日期": date_text,
                "总人数": sum(status_counter.values()),
                **status_counter,
            }
        )

    summary_rows = []
    for key, dates in absent_dates.items():
        info = student_info[key]
        summary_rows.append(
            {
                **info,
                "缺勤次数": len(dates),
                "缺勤日期": "、".join(sorted(dates)),
            }
        )
    summary_rows.sort(key=lambda row: (-row["缺勤次数"], str(row["姓名"])))
    detail_rows.sort(key=lambda row: (row["日期"], str(row["姓名"])))
    overview_rows.sort(key=lambda row: row["日期"])

    write_results(len(excel_files), summary_rows, detail_rows, overview_rows)

    print(f"已统计 {len(excel_files)} 次签到。")
    print(f"有缺勤记录学生数：{len(summary_rows)}")
    print(f"缺勤记录总条数：{len(detail_rows)}")
    print(f"结果已保存到：{OUTPUT_FILE}")


if __name__ == "__main__":
    main()
